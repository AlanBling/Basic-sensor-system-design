#!/usr/local/bin/python

# Author: Yuxuan Zhang
# Initial date: 12.July.2021
# Project name: Basic sensor system with UI 

# Use thread
import threading

# Operate the Excel file
from geoip2.types import IPAddress
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd

# Create the time stamp
from datetime import datetime

# Make time delay
import time

# Use to control DHT22
import Adafruit_DHT as DHT

# Use to contorl the GPIO on the Raspberry Pi board
import RPi.GPIO as GPIO

# Use to detect the IP address
from urllib import request
import requests

# Use to detect the Physical address
import geoip2.database

# Use to read the HTML file
from lxml import etree

# Import the PyQt5 moddule
from PyQt5.QtCore import QThread, Qt, pyqtSignal
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QGridLayout
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# Import to make the figure
from pylab import mpl
import matplotlib.pyplot as plt
import numpy as np
import sys
import matplotlib

# Setting the parameter
matplotlib.use('Qt5Agg')

mpl.rcParams['axes.unicode_minus'] = False

# Ignore the GPIO from the system
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BOARD)

#define the pin that goes to the circuit
photoresistor_PIN = 37

# Setting DHT_Sensor model 
DHT_Sensor = 22

# Setting DHT_Sensor PIN number
DHT_Sensor_PIN = 4

# Setting the red LED as one of appliances, the red LED PIN is 21
redLED_PIN = 40
GPIO.setup(redLED_PIN, GPIO.OUT)

# Setting the red LED working threshold value
thresholdForHumidity = 60

# Create the time stamp
ticks = str(datetime.now())[:19]
# Clear the time stamp, use "blankspace" to substitute the ",",":","-"
recordingFileName = str(ticks).replace('-','').replace(':','').replace(' ','')

# Create a new Excel file
recordingFile = Workbook()

# # Count the raw number to create the new data into the Excel file
numOfRaw = 2

# Global variable
IPAddress = ""
countryName = ""
cityName = ""
postcode = ""
timezone = ""
timeStamp = ""

valueOfHumidity = 0
valueOfTemperature = 0
valueOfLightIntensity = 0

weatherInformation = ""

previousTime = 0

# Locate to the active sheet
recordingFile_Sheet1 = recordingFile.active

# Create the table heading
recordingFile_Sheet1['A1'] = "Time Stamp"
recordingFile_Sheet1['B1'] = "Temperature"
recordingFile_Sheet1['C1'] = "Humidity"
recordingFile_Sheet1['D1'] = "Light Intensity"
recordingFile_Sheet1['E1'] = "IPAddress"
recordingFile_Sheet1['F1'] = "countryName"
recordingFile_Sheet1['G1'] = "cityName"
recordingFile_Sheet1['H1'] = "postcode"
recordingFile_Sheet1['I1'] = "timezone"
recordingFile_Sheet1['J1'] = "Weather"


# Save the Excel file
recordingFile.save("./" + recordingFileName +".xlsx")

# photoresistor circuit detection
def rc_time (photoresistor_PIN):

    count = 0
  
    GPIO.setup(photoresistor_PIN, GPIO.OUT)
    GPIO.output(photoresistor_PIN, GPIO.LOW)
    time.sleep(0.1)

    GPIO.setup(photoresistor_PIN, GPIO.IN)
  
    while (GPIO.input(photoresistor_PIN) == GPIO.LOW):
        count += 1

    return count

# Detect the IP address via the external website
def detectIPAdress():
    return request.urlopen("https://ip.42.pl/raw").read().decode("utf8", errors="ignore")


def detectLocation():

    # Generate the location reader
    reader = geoip2.database.Reader("./lib/GeoLite2-City.mmdb")
    
    # Call IP detection function
    ip = detectIPAdress()
    IPAddress = ip

    # print(type(IPAddress))
    
    # Get response from "Location reader"
    response = reader.city(ip)

    # Get "Country, City, Postcode, Timezone"
    countryName = response.country.name
    cityName = response.city.name
    postcode = response.postal.code
    timezone = response.location.time_zone

    return (IPAddress, countryName, cityName, postcode, timezone)

# Detect the user location first time
IPAddress, countryName, cityName, postcode, timezone = detectLocation()

def weatherInfo():
    url1 = "https://www.weather-forecast.com/locations/"
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.02311.135 Safari/537.36 Edge/12.10240"
    url2 = "/forecasts/latest"
    
    # print(cityName)
    with requests.request("GET", url1 + cityName + url2, headers={"User-agent":ua}) as res:
        content = res.text
        html = etree.HTML(content)
        weatherInfo = html.xpath("/html/body/main/section[3]/div/div/div[2]/div/table/thead/tr[1]/td[1]/p/span/text()")
    try:
        return weatherInfo[0]
    except:
        return "Oops, looks like some errors occured. Please check your Internet connection."

# Inquiry the weather information when the system initialised.
weatherInformation = weatherInfo()

# Sensor system core section
class sensorThread (QThread):

    SIGvalueOfHumidity = pyqtSignal(int)
    SIGvalueOfTemperature = pyqtSignal(int)
    SIGvalueOfLightIntensity = pyqtSignal(int)
    SIGtimeStamp = pyqtSignal(str)

    # Initial function
    def __init__(self, threadID, name):
        super().__init__()
        # threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name

    def run(self) -> None:

        global numOfRaw, previousTime, IPAddress, countryName, cityName, postcode, timezone, weatherInformation, valueOfHumidity, valueOfTemperature, valueOfLightIntensity, timeStamp

        ##############################
        ###Sensor system MAIN BODY ###
        while True:
            
            # Read value from sensor, store the value into the variable
            valueOfHumidity, valueOfTemperature = DHT.read_retry(DHT_Sensor, DHT_Sensor_PIN)
            
            # The value from photoresistance circuit
            valueOfLightIntensity = rc_time(photoresistor_PIN)
            
            if valueOfHumidity != None and valueOfHumidity != None:

                # Update the timeStamp
                timeStamp = str(datetime.now())[:19]

                currentTime = int(timeStamp[11:13])
                if currentTime > previousTime or currentTime == 1:
                    previousTime  = currentTime
                    weatherInformation = weatherInfo()
                
                self.SIGtimeStamp.emit(timeStamp)
                self.SIGvalueOfHumidity.emit(valueOfHumidity)
                self.SIGvalueOfTemperature.emit(valueOfTemperature)
                self.SIGvalueOfLightIntensity.emit(valueOfLightIntensity)

                # Update the data recording
                recordingFile = load_workbook("./"+recordingFileName+".xlsx")
                
                # Locate the active sheet
                Sheet1 = recordingFile["Sheet"]
                # Update the each columns
                Sheet1['A'+str(numOfRaw)] = timeStamp
                Sheet1['B'+str(numOfRaw)] = valueOfTemperature
                Sheet1['C'+str(numOfRaw)] = valueOfHumidity
                Sheet1['D'+str(numOfRaw)] = valueOfLightIntensity
                Sheet1['E'+str(numOfRaw)] = IPAddress
                Sheet1['F'+str(numOfRaw)] = countryName
                Sheet1['G'+str(numOfRaw)] = cityName
                Sheet1['H'+str(numOfRaw)] = postcode
                Sheet1['I'+str(numOfRaw)] = timezone
                Sheet1['J'+str(numOfRaw)] = weatherInformation

                # Move to the next row
                numOfRaw += 1

                # Save the excel file avoid the progrmme interrupt by unpredict errors
                recordingFile.save("./"+recordingFileName+".xlsx")

            else:
                # Came across the error
                print("DHT22 get value failed, check wires please.")
                continue
            

            # Setting the working rule for red LED
            if valueOfHumidity > thresholdForHumidity:                
                GPIO.output(redLED_PIN, GPIO.HIGH)
            else:
                GPIO.output(redLED_PIN, GPIO.LOW)

            # Delay for next time read value from sensors
            # If make this block working, then the system working frequency will be impact.
            # time.sleep(0.5)

# System UI section
class App(QWidget):

    # global valueOfHumidity, valueOfTemperature, valueOfLightIntensity, timeStamp
    valueOfHumidity = 0
    valueOfTemperature = 0
    valueOfLightIntensity = 0
    timeStamp = ""

    def __init__(self, parent=None):
        # Initialize the father class function.
        super(App, self).__init__(parent)

        self.flag_temperaturePermission = True
        self.flag_humidityPermission = True
        self.flag_lightintensityPermission = True

        self.secondsSlectionResult = 1
        self.minutesSlectionResult = 0
        self.hoursSelectionResult = 0

        self.performanceFanStep = 100
        self.performanceHumidifierStep = 100
        self.performanceCurtainStep = 100

        self.update()

        self.initUI()

    def update(self):
        # Create a thread
        self.current = sensorThread("Thread-1", "Basic sensor system")
        # Runing the socket function after recieved the signal
        self.current.SIGtimeStamp.connect(self.GetTimeStamp)
        self.current.SIGvalueOfTemperature.connect(self.GetTemperature)
        self.current.SIGvalueOfHumidity.connect(self.GetHumidity)
        self.current.SIGvalueOfLightIntensity.connect(self.GetLightIntensity)

        self.current.start()

    def initUI(self):

        self.setWindowTitle('Sensor system demo application')
        self.resize(1200, 750)

        # Set the window at the centre of the screen
        self.center()

        # Some QWidgets
        self.temperaturePermissionButton = QPushButton()
        # Locating the pic URL into the button.
        self.temperaturePermissionButton.setIcon(QIcon("./resource/temperature_icon.png")) 
        # Setting the picture size in the button.
        self.temperaturePermissionButton.setIconSize(QtCore.QSize(96, 96))

        self.humidityPermissionButton = QPushButton()
        self.humidityPermissionButton.setIcon(QIcon("./resource/humidity_icon.png"))
        self.humidityPermissionButton.setIconSize(QtCore.QSize(96, 96))

        self.lightintensityPermissionButton = QPushButton()
        self.lightintensityPermissionButton.setIcon(QIcon("./resource/lightintensity.png"))
        self.lightintensityPermissionButton.setIconSize(QtCore.QSize(96, 96))

        self.temperaturePermissionButton.clicked.connect(self.temperaturePermissionControl)
        self.humidityPermissionButton.clicked.connect(self.humidityPermissionControl)
        self.lightintensityPermissionButton.clicked.connect(self.lightintensityPermissionControl)


        self.label_TemperatureValue = QLabel()
        self.label_TemperatureValue.setStyleSheet('color:red; font:bold 24px;')
        self.label_HumidityValue = QLabel()
        self.label_HumidityValue.setStyleSheet('color:blue; font:bold 24px;')
        self.label_LightIntensityValue = QLabel()
        self.label_LightIntensityValue.setStyleSheet('color:orange; font:bold 24px;')


        fanPic = QPixmap('./resource/fan_icon.png')
        self.serviceThresholdIcon_Fan = QLabel()
        self.serviceThresholdIcon_Fan.setPixmap(fanPic)
        self.serviceThresholdIcon_Fan.setScaledContents(True)
        self.serviceThresholdIcon_Fan.setMaximumSize (90, 90)

        self.threshold_fan = QLineEdit("33")
        self.unit_fan = QLabel("°C")

        humidifierPic = QPixmap('./resource/humidifier_icon.png')
        self.serviceThresholdIcon_Humidifier = QLabel()
        self.serviceThresholdIcon_Humidifier.setPixmap(humidifierPic)
        self.serviceThresholdIcon_Humidifier.setScaledContents(True)
        self.serviceThresholdIcon_Humidifier.setMaximumSize(90, 90)

        self.threshold_humidifier = QLineEdit("30")
        self.unit_humidifier = QLabel("%RH")

        curtainPic = QPixmap('./resource/curtain_icon.png')
        self.serviceThresholdIcon_Curtain = QLabel()
        self.serviceThresholdIcon_Curtain.setPixmap(curtainPic)
        self.serviceThresholdIcon_Curtain.setScaledContents(True)
        self.serviceThresholdIcon_Curtain.setMaximumSize(90, 90)

        self.threshold_curtain = QLineEdit("Low")
        self.unit_curtain = QLabel("Level")

        self.label_time_scale = QLabel("Time scale")

        self.button_reset_timescale = QPushButton("Reset")
        self.button_reset_timescale.clicked.connect(self.resetTimeScale)

        # Service icon settings
        fanPic = QPixmap('./resource/fan_icon.png')
        self.serviceIcon_Fan = QLabel()
        self.serviceIcon_Fan.setStyleSheet("border: 10px solid red")
        self.serviceIcon_Fan.setPixmap(fanPic)
        self.serviceIcon_Fan.setScaledContents(True)
        self.serviceIcon_Fan.setMaximumSize (120, 120)

        humidifierPic = QPixmap('./resource/humidifier_icon.png')
        self.serviceIcon_Humidifier = QLabel()
        self.serviceIcon_Humidifier.setStyleSheet("border: 10px solid blue")
        self.serviceIcon_Humidifier.setPixmap(humidifierPic)
        self.serviceIcon_Humidifier.setScaledContents(True)
        self.serviceIcon_Humidifier.setMaximumSize (120, 120)

        curtainPic = QPixmap('./resource/curtain_icon.png')
        self.serviceIcon_Curtain = QLabel()
        self.serviceIcon_Curtain.setStyleSheet("border: 10px solid orange")
        self.serviceIcon_Curtain.setPixmap(curtainPic)
        self.serviceIcon_Curtain.setScaledContents(True)
        self.serviceIcon_Curtain.setMaximumSize (120, 120)

        # Label of system status INFO
        label_systemStatusINFO = QLabel("System status INFO")
        label_systemStatusINFO.setStyleSheet("font:bold 25px;")

        label_running_status = QLabel("Running status")
        label_running_status.setStyleSheet("font:bold 20px;")

        label_service_fan = QLabel("Fan")
        label_service_fan.setStyleSheet("font:15px;")
        label_service_humidifier = QLabel("Humidifier")
        label_service_humidifier.setStyleSheet("font:15px;")
        label_service_curtain = QLabel("Curtain")
        label_service_curtain.setStyleSheet("font:15px;")

        self.label_status_fan = QLabel("Automatically")
        self.label_status_fan.setStyleSheet("font:15px;")
        self.label_status_humidifier = QLabel("Automatically")
        self.label_status_humidifier.setStyleSheet("font:15px;")
        self.label_status_curtain = QLabel("Automatically")
        self.label_status_curtain.setStyleSheet("font:15px;")

        self.label_service_performance = QLabel("Service performance")
        self.label_service_performance.setStyleSheet("font:bold 20px;")

        # Progress bar group
        self.bar_fan = QProgressBar()
        # The directioin of the progress bar
        self.bar_fan.setInvertedAppearance(False)
        self.bar_fan.setOrientation(Qt.Vertical)
        self.bar_fan.setMaximum(100)

        self.bar_humidifier = QProgressBar()
        # The directioin of the progress bar
        self.bar_humidifier.setInvertedAppearance(False)
        self.bar_humidifier.setOrientation(Qt.Vertical)
        self.bar_humidifier.setMaximum(100)

        
        self.bar_curtain = QProgressBar()
        # The directioin of the progress bar
        self.bar_curtain.setInvertedAppearance(False)
        self.bar_curtain.setOrientation(Qt.Vertical)
        self.bar_curtain.setMaximum(100)

        # Initialize the QComBox object
        self.comboBox_seconds = QComboBox()
        self.comboBox_seconds.setMaximumWidth(45)
        # Add several items
        list_1_30 = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30"]
        self.comboBox_seconds.addItems(list_1_30)
        
        # When the comboxs' content changed running the pointed event.
        self.comboBox_seconds.currentIndexChanged.connect(self.secondSelection)
        self.secondsLabel = QLabel("Seconds")
        self.secondsLabel.setMaximumWidth(50)

        # Initialize the QComBox object
        self.comboBox_minutes = QComboBox()
        self.comboBox_minutes.setMaximumWidth(45)
        # Add several items
        list_0_30 = ["0","1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30"]
        self.comboBox_minutes.addItems(list_0_30)
        # When the comboxs' content changed running the pointed event.
        self.comboBox_minutes.currentIndexChanged.connect(self.minuteSelection)
        self.minutesLabel = QLabel("Minutes")
        self.minutesLabel.setMaximumWidth(50)

        # Initialize the QComBox object
        self.comboBox_hours = QComboBox()
        self.comboBox_hours.setMaximumWidth(45)
        # Add several items
        list_0_5 = ["0","1","2","3","4","5"]
        # self.comboBox_hours.addItems(["0","1","2","3","4","5"])
        self.comboBox_hours.addItems(list_0_5)
        # When the comboxs' content changed running the pointed event.
        self.comboBox_hours.currentIndexChanged.connect(self.hourSelection)
        self.hoursLabel = QLabel("Hours")
        self.hoursLabel.setMaximumWidth(40)
        
        self.startBtn = QPushButton('Start')
        self.endBtn = QPushButton('Over')

        self.startBtn.clicked.connect(self.startTimer)
        self.endBtn.clicked.connect(self.endTimer)

        # Time module
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.showTime)

        # Figure module
        plt.xticks(rotation=45)
        self.figure, self.ax = plt.subplots()
        
        self.ax2 = self.ax.twinx()

        self.canvas = FigureCanvas(self.figure)

        # Box layout settings
        startOverButtonsBox = QHBoxLayout()
        startOverButtonsBox.addWidget(self.startBtn)
        startOverButtonsBox.addWidget(self.endBtn)

        permissionButtonsBox = QHBoxLayout()
        permissionButtonsBox.addWidget(self.temperaturePermissionButton)
        permissionButtonsBox.addWidget(self.humidityPermissionButton)
        permissionButtonsBox.addWidget(self.lightintensityPermissionButton)

        label_ValueBox = QHBoxLayout()
        label_ValueBox.addWidget(self.label_TemperatureValue)
        label_ValueBox.addWidget(self.label_HumidityValue)
        label_ValueBox.addWidget(self.label_LightIntensityValue)

        label_ServiceBox = QHBoxLayout()
        label_ServiceBox.addWidget(self.serviceIcon_Fan)
        label_ServiceBox.addWidget(self.serviceIcon_Humidifier)
        label_ServiceBox.addWidget(self.serviceIcon_Curtain)

        threshold_fan_Box = QHBoxLayout()
        threshold_fan_Box.addWidget(self.serviceThresholdIcon_Fan)
        threshold_fan_Box.addWidget(self.threshold_fan)
        threshold_fan_Box.addWidget(self.unit_fan)

        threshold_humidifier_Box = QHBoxLayout()
        threshold_humidifier_Box.addWidget(self.serviceThresholdIcon_Humidifier)
        threshold_humidifier_Box.addWidget(self.threshold_humidifier)
        threshold_humidifier_Box.addWidget(self.unit_humidifier)

        threshold_curtain_Box = QHBoxLayout()
        threshold_curtain_Box.addWidget(self.serviceThresholdIcon_Curtain)
        threshold_curtain_Box.addWidget(self.threshold_curtain)
        threshold_curtain_Box.addWidget(self.unit_curtain)

        threshold_Box = QVBoxLayout()
        threshold_Box.addLayout(threshold_fan_Box)
        threshold_Box.addLayout(threshold_humidifier_Box)
        threshold_Box.addLayout(threshold_curtain_Box)

        permission_service_Box = QVBoxLayout()
        permission_service_Box.addLayout(startOverButtonsBox)
        permission_service_Box.addLayout(permissionButtonsBox)
        permission_service_Box.addLayout(label_ValueBox)
        permission_service_Box.addLayout(label_ServiceBox)

        argument_permission_Box = QHBoxLayout()
        argument_permission_Box.addLayout(threshold_Box)
        argument_permission_Box.addLayout(permission_service_Box)

        label_service_Box = QHBoxLayout()
        label_service_Box.addWidget(label_service_fan)
        label_service_Box.addWidget(label_service_humidifier)
        label_service_Box.addWidget(label_service_curtain)
        
        service_status_Box = QHBoxLayout()
        service_status_Box.addWidget(self.label_status_fan)
        service_status_Box.addWidget(self.label_status_humidifier)
        service_status_Box.addWidget(self.label_status_curtain)

        progress_bar_Box = QHBoxLayout()
        progress_bar_Box.addWidget(self.bar_fan)
        progress_bar_Box.addWidget(self.bar_humidifier)
        progress_bar_Box.addWidget(self.bar_curtain)

        system_Status_INFO_Box = QVBoxLayout()
        system_Status_INFO_Box.addWidget(label_systemStatusINFO)
        system_Status_INFO_Box.addWidget(label_running_status)
        system_Status_INFO_Box.addLayout(label_service_Box)
        system_Status_INFO_Box.addLayout(service_status_Box)
        system_Status_INFO_Box.addWidget(self.label_service_performance)
        system_Status_INFO_Box.addLayout(progress_bar_Box)

        upper_Box = QHBoxLayout()
        upper_Box.addStretch(1)
        upper_Box.addLayout(argument_permission_Box)
        upper_Box.addStretch(1)
        upper_Box.addLayout(system_Status_INFO_Box)
        upper_Box.addStretch(2)

        time_scale_comobox_Box = QHBoxLayout()
        time_scale_comobox_Box.addWidget(self.comboBox_seconds)
        time_scale_comobox_Box.addWidget(self.secondsLabel)
        time_scale_comobox_Box.addWidget(self.comboBox_minutes)
        time_scale_comobox_Box.addWidget(self.minutesLabel)
        time_scale_comobox_Box.addWidget(self.comboBox_hours)
        time_scale_comobox_Box.addWidget(self.hoursLabel)

        time_scale_Box = QVBoxLayout()
        time_scale_Box.addStretch(1)
        time_scale_Box.addLayout(time_scale_comobox_Box)
        time_scale_Box.addWidget(self.label_time_scale)
        time_scale_Box.addWidget(self.button_reset_timescale)
        time_scale_Box.addStretch(1)

        timeControl_Chart_Box = QHBoxLayout()
        timeControl_Chart_Box.addStretch(1)
        timeControl_Chart_Box.addLayout(time_scale_Box)
        timeControl_Chart_Box.addWidget(self.canvas)
        timeControl_Chart_Box.addStretch(1)

        mainViewBox = QVBoxLayout()
        mainViewBox.addStretch(1)
        mainViewBox.addLayout(upper_Box)
        mainViewBox.addStretch(1)
        mainViewBox.addLayout(timeControl_Chart_Box)
        
        self.setLayout(mainViewBox)

        # Initialize the arrays.
        self.x = []
        self.y = []
        self.z = []
        self.timeIndex = []

    def center(self):
        qr = self.frameGeometry()
        # Get the size of the window.
        print('qr:', qr)
        cp = QDesktopWidget().availableGeometry().center()
        # Get the distinguishability value, get the cneter point location.
        print('cp:', cp)
        qr.moveCenter(cp)
        # Make the window's centre point at the QR center point
        self.move(qr.topLeft())

    def showTime(self):

        # Update the arrays.
        if len(self.timeIndex) == 10:
            if timeStamp[11:] not in self.timeIndex and timeStamp[11:13] not in self.timeIndex and timeStamp[11:16] not in self.timeIndex:        

                self.timeIndex = self.timeIndex[1:]

                if self.hoursSelectionResult != 0:
                    self.timeIndex.append(timeStamp[11:13])
                elif self.minutesSlectionResult != 0:
                    self.timeIndex.append(timeStamp[11:16])
                else:
                    self.timeIndex.append(timeStamp[11:])
                
                if self.flag_humidityPermission:
                    self.x = self.x[1:]
                    self.x.append(round(valueOfHumidity, 1))
                else:
                    self.x = self.x[1:]
                    self.x.append(-100)
                
                if self.flag_temperaturePermission:
                    self.y = self.y[1:]
                    self.y.append(round(valueOfTemperature, 1))
                else:
                    self.y = self.y[1:]
                    self.y.append(-100)
                
                if self.flag_lightintensityPermission:
                    self.z = self.z[1:]
                    if valueOfLightIntensity > 40000:
                        self.z.append(22)
                    else:
                        self.z.append(32)
                else:
                    self.z = self.z[1:]
                    self.z.append(-100)

        else:
            if timeStamp[11:] not in self.timeIndex and timeStamp[11:13] not in self.timeIndex and timeStamp[11:16] not in self.timeIndex:

                if self.hoursSelectionResult != 0:
                    self.timeIndex.append(timeStamp[11:13])
                elif self.minutesSlectionResult != 0:
                    self.timeIndex.append(timeStamp[11:16])
                else:
                    self.timeIndex.append(timeStamp[11:])

                if self.flag_humidityPermission:
                    self.x.append(round(valueOfHumidity, 1))
                else:
                    self.x.append(-100)

                if self.flag_temperaturePermission:
                    self.y.append(round(valueOfTemperature, 1))
                else:
                    self.y.append(-100)
                
                if self.flag_lightintensityPermission:
                    if valueOfLightIntensity > 40000:
                        self.z.append(22)
                    else:
                        self.z.append(32)
                else:
                    self.z.append(-100)
        
        self.ax.clear()
        self.ax2.clear()

        self.ax.plot(self.timeIndex, self.x, color="blue", label="Humidity")
        self.ax.set_ylabel("Humidity RH%")
        self.ax.set_xlabel("Time")
        
        self.ax.set_ylim(0,100)

        self.ax2.plot(self.timeIndex, self.y, color = "red", label="Temperature")
        self.ax2.plot(self.timeIndex, self.z, color = "orange", label="Light intensity")
        self.ax2.set_ylabel("Temperature °C")

        if self.y[-1] != -100:
            self.label_TemperatureValue.setText(str(self.y[-1]) + " °C")
            self.serviceIcon_Fan.setStyleSheet("border: 10px solid red")
            self.label_status_fan.setText("Automatically")
            
            if self.hoursSelectionResult > 0:
                self.performanceFanStep = 25 - self.hoursSelectionResult * 5
            elif self.minutesSlectionResult > 0:
                self.performanceFanStep = 75 - self.minutesSlectionResult * 1.66
            else:
                self.performanceFanStep = 100 - self.secondsSlectionResult * 1.2

        else:
            self.label_TemperatureValue.setText("Unknown")
            self.serviceIcon_Fan.setStyleSheet("border: 10px solid grey")
            self.label_status_fan.setText("Manual")

            self.performanceFanStep = 0

        if self.x[-1] != -100:
            self.label_HumidityValue.setText(str(self.x[-1]) + " %RH")
            self.serviceIcon_Humidifier.setStyleSheet("border: 10px solid blue")
            self.label_status_humidifier.setText("Automatically")

            if self.hoursSelectionResult > 0:
                self.performanceHumidifierStep = 50 - self.hoursSelectionResult * 10
            elif self.minutesSlectionResult > 0:
                self.performanceHumidifierStep = 75 - self.minutesSlectionResult * 0.833
            else:
                self.performanceHumidifierStep = 100 - self.secondsSlectionResult * 0.833

        else:
            self.label_HumidityValue.setText("Unknown")
            self.serviceIcon_Humidifier.setStyleSheet("border: 10px solid grey")
            self.label_status_humidifier.setText("Manual")

            self.performanceHumidifierStep = 0
        
        if self.z[-1] != -100:
            if self.z[-1] == 22:
                self.label_LightIntensityValue.setText("Low")
            else:
                self.label_LightIntensityValue.setText("High")
            self.serviceIcon_Curtain.setStyleSheet("border: 10px solid orange")
            self.label_status_curtain.setText("Automatically")
            
            if self.hoursSelectionResult > 0:
                self.performanceCurtainStep = 85 - self.hoursSelectionResult * 17
            elif self.minutesSlectionResult > 0:
                self.performanceCurtainStep = 90 - self.minutesSlectionResult * 0.1666
            else:
                self.performanceCurtainStep = 100 - self.secondsSlectionResult * 0.333

        else:
            self.label_LightIntensityValue.setText("Unknown")
            self.serviceIcon_Curtain.setStyleSheet("border: 10px solid grey")
            self.label_status_curtain.setText("Manual")
            self.performanceCurtainStep = 0

        self.figure.legend(loc="upper left", bbox_to_anchor=(0, 1), bbox_transform=self.ax.transAxes)

        self.bar_fan.setValue(self.performanceFanStep)
        self.bar_humidifier.setValue(self.performanceHumidifierStep)
        self.bar_curtain.setValue(self.performanceCurtainStep)

        if self.performanceFanStep > 79:
            self.bar_fan.setStyleSheet("QProgressBar::chunk{background-color: #00ff00}")
        elif self.performanceFanStep > 39:
            self.bar_fan.setStyleSheet("QProgressBar::chunk{background-color: #f0b21d}")
        else:
            self.bar_fan.setStyleSheet("QProgressBar::chunk{background-color: #ea380b}")

        if self.performanceHumidifierStep > 79:
            self.bar_humidifier.setStyleSheet("QProgressBar::chunk{background-color: #00ff00}")
        elif self.performanceHumidifierStep > 39:
            self.bar_humidifier.setStyleSheet("QProgressBar::chunk{background-color: #f0b21d}")
        else:
            self.bar_humidifier.setStyleSheet("QProgressBar::chunk{background-color: #ea380b}")

        if self.performanceCurtainStep > 79:
            self.bar_curtain.setStyleSheet("QProgressBar::chunk{background-color: #00ff00}")
        elif self.performanceCurtainStep > 39:
            self.bar_curtain.setStyleSheet("QProgressBar::chunk{background-color: #f0b21d}")
        else:
            self.bar_curtain.setStyleSheet("QProgressBar::chunk{background-color: #ea380b}")
    
        self.ax2.set_ylim(20,35)

        plt.subplots_adjust(bottom=0.2)

        for tick in self.ax.get_xticklabels():
            tick.set_rotation(45)

        self.canvas.draw()

    # Start up function
    def startTimer(self):
        # Setting the timer and make the system running
        self.timer.start(1)
        # Forbidden the "Start" button
        self.startBtn.setEnabled(False)
        #  Make the "Over" button useable.
        self.endBtn.setEnabled(True) 

    def endTimer(self):
        # Stop the timer
        self.timer.stop()
        # Forbidden the "Start" button
        self.startBtn.setEnabled(True)
        #  Make the "Over" button useable.
        self.endBtn.setEnabled(False)
        # Clear the arrays
        self.x = []
        self.y = []
        self.timeIndex = []
        self.z = []

    def temperaturePermissionControl(self):
        if self.flag_temperaturePermission:
            self.flag_temperaturePermission = False
        else:
            self.flag_temperaturePermission = True

    def humidityPermissionControl(self):
        if self.flag_humidityPermission:
            self.flag_humidityPermission = False
        else:
            self.flag_humidityPermission = True

    def lightintensityPermissionControl(self):
        if self.flag_lightintensityPermission:
            self.flag_lightintensityPermission = False
        else:
            self.flag_lightintensityPermission = True

    def secondSelection(self):
        # The lable use to show the text.
        self.secondsSlectionResult = int(self.comboBox_seconds.currentText())

    def minuteSelection(self):
        self.minutesSlectionResult = int(self.comboBox_minutes.currentText())

    def hourSelection(self):
        self.hoursSelectionResult = int(self.comboBox_hours.currentText())

    def GetTimeStamp(self, signal):
        self.timeStamp = signal

    def GetTemperature(self, signal):
        self.valueOfTemperature = signal

    def GetHumidity(self, signal):
        self.valueOfHumidity = signal

    def GetLightIntensity(self, signal):
        self.valueOfLightIntensity = signal

    def resetTimeScale(self):
        self.comboBox_hours.setCurrentIndex(0)
        self.comboBox_minutes.setCurrentIndex(0)
        self.comboBox_seconds.setCurrentIndex(0)

class UIThread(threading.Thread):
    # Initial function
    def __init__(self, threadID, name):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name

    def run(self) -> None:
        
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
        
        app = QApplication(sys.argv)

        main_window = App()
        main_window.show()

        app.exec()


# Main programe enter
if __name__ == '__main__':
    userInterfaceThread = UIThread("Thread-2", "UI")
    userInterfaceThread.start()
    userInterfaceThread.join()