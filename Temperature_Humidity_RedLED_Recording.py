#!/usr/local/bin/python

# Author: Alan Zhang
# Initial date: 12.July.2021
# Project name: Basic sensor system


# Operate the Excel file
from geoip2.types import IPAddress
from openpyxl import Workbook
from openpyxl import load_workbook

# Create the time stamp
from datetime import datetime

# Make time delay
import time

# Use to control DHT22
import Adafruit_DHT as DHT

import RPi.GPIO as GPIO

# Use to detect the IP address
from urllib import request
import requests

# Use to detect the Physical address
import geoip2.database

# Use to read the HTML file
from lxml import etree


GPIO.setwarnings(False)
GPIO.setmode(GPIO.BOARD)

#define the pin that goes to the circuit
photoresistor_PIN = 37


# Setting DHT_Sensor model 
DHT_Sensor = 22

# Setting DHT_Sensor PIN number
DHT_Sensor_PIN = 4

# Setting the red LED as one of appliances, the red LED PIN is 21
redLED_PIN = 40
GPIO.setup(redLED_PIN, GPIO.OUT)

# Setting the red LED working threshold value
thresholdForHumidity = 60

# Create the time stamp
ticks = str(datetime.now())[:19]
# Clear the time stamp, use "blankspace" to substitute the ",",":","-"
recordingFileName = str(ticks).replace('-','').replace(':','').replace(' ','')

# Create a new Excel file
recordingFile = Workbook()

# Count the raw number to create the new data into the Excel file
numOfRaw = 2

# Global variable
IPAddress = ""
countryName = ""
cityName = ""
postcode = ""
timezone = ""

weatherInformation = ""

previousTime = 0

# Locate to the active sheet
recordingFile_Sheet1 = recordingFile.active

# Create the table heading
recordingFile_Sheet1['A1'] = "Time Stamp"
recordingFile_Sheet1['B1'] = "Temperature"
recordingFile_Sheet1['C1'] = "Humidity"
recordingFile_Sheet1['D1'] = "Light Intensity"
recordingFile_Sheet1['E1'] = "IPAddress"
recordingFile_Sheet1['F1'] = "countryName"
recordingFile_Sheet1['G1'] = "cityName"
recordingFile_Sheet1['H1'] = "postcode"
recordingFile_Sheet1['I1'] = "timezone"
recordingFile_Sheet1['J1'] = "Weather"


# Save the Excel file
recordingFile.save("./" + recordingFileName +".xlsx")



# photoresistor circuit detection
def rc_time (photoresistor_PIN):

    count = 0
  
    GPIO.setup(photoresistor_PIN, GPIO.OUT)
    GPIO.output(photoresistor_PIN, GPIO.LOW)
    time.sleep(0.1)

    GPIO.setup(photoresistor_PIN, GPIO.IN)
  
    while (GPIO.input(photoresistor_PIN) == GPIO.LOW):
        count += 1

    return count


# Detect the IP address via the external website
def detectIPAdress():
    return request.urlopen("https://ip.42.pl/raw").read().decode("utf8", errors="ignore")



def detectLocation():

    # Generate the location reader
    reader = geoip2.database.Reader("./lib/GeoLite2-City.mmdb")
    
    # Call IP detection function
    ip = detectIPAdress()
    IPAddress = ip

    # print(type(IPAddress))
    
    # Get response from "Location reader"
    response = reader.city(ip)


    # Get "Country, City, Postcode, Timezone"
    countryName = response.country.name
    cityName = response.city.name
    postcode = response.postal.code
    timezone = response.location.time_zone

    return (IPAddress, countryName, cityName, postcode, timezone)


# Detect the user location first time
IPAddress, countryName, cityName, postcode, timezone = detectLocation()


def weatherInfo():
    url1 = "https://www.weather-forecast.com/locations/"
    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.02311.135 Safari/537.36 Edge/12.10240"
    url2 = "/forecasts/latest"
    
    # print(cityName)
    with requests.request("GET", url1 + cityName + url2, headers={"User-agent":ua}) as res:
        content = res.text
        html = etree.HTML(content)
        weatherInfo = html.xpath("/html/body/main/section[3]/div/div/div[2]/div/table/thead/tr[1]/td[1]/p/span/text()")
    try:
        return weatherInfo[0]
    except:
        return "Oops, looks like some errors occured. Please check your Internet connection."

# Inquiry the weather information when the system initialised.
weatherInformation = weatherInfo()

#################
### MAIN BODY ###
while True:
    
    # Read value from sensor, store the value into the variable
    valueOfHumidity, valueOfTemperature = DHT.read_retry(DHT_Sensor, DHT_Sensor_PIN)

    # The value from photoresistance circuit
    valueOfLightIntensity = rc_time(photoresistor_PIN)

    if numOfRaw % 7 == 0:
    
        # Detect the user location
        IPAddress, countryName, cityName, postcode, timezone = detectLocation()
    
    if valueOfHumidity != None and valueOfHumidity != None:

        # Update the timeStamp
        timeStamp = str(datetime.now())[:19]

        currentTime = int(timeStamp[11:13])
        if currentTime > previousTime or currentTime == 1:
            previousTime  = currentTime
            weatherInformation = weatherInfo()

    
        # Output the value of temperature and humidity
        print('%s, Temperature = %.1f *C, Humidity = %.1f%%RH, Light intensity = %d, Country = %s, City = %s, Postcode = %s, Timezone = %s, IP = %s\n Weather forecast: %s\n' % (timeStamp, valueOfTemperature, valueOfHumidity, valueOfLightIntensity, countryName, cityName, postcode, timezone, IPAddress, weatherInformation))
        
        # Update the data recording
        recordingFile = load_workbook("./"+recordingFileName+".xlsx")
        
        # Locate the active sheet
        Sheet1 = recordingFile["Sheet"]
        # Update the each columns
        Sheet1['A'+str(numOfRaw)] = timeStamp
        Sheet1['B'+str(numOfRaw)] = valueOfTemperature
        Sheet1['C'+str(numOfRaw)] = valueOfHumidity
        Sheet1['D'+str(numOfRaw)] = valueOfLightIntensity
        Sheet1['E'+str(numOfRaw)] = IPAddress
        Sheet1['F'+str(numOfRaw)] = countryName
        Sheet1['G'+str(numOfRaw)] = cityName
        Sheet1['H'+str(numOfRaw)] = postcode
        Sheet1['I'+str(numOfRaw)] = timezone
        Sheet1['J'+str(numOfRaw)] = weatherInformation

        # Move to the next row
        numOfRaw += 1

        # Save the excel file avoid the progrmme interrupt by unpredict errors
        recordingFile.save("./"+recordingFileName+".xlsx")

    else:
        # Came across error
        print("DHT22 get value failed, check wires please.")
        continue
    

    # Setting the working rule for red LED
    if valueOfHumidity > thresholdForHumidity:
        
        GPIO.output(redLED_PIN, GPIO.HIGH)

    else:

        GPIO.output(redLED_PIN, GPIO.LOW)

    # Delay for next time read value from sensors
    time.sleep(2.5)







